from dagster import (
    asset,
    AssetExecutionContext,
    Definitions,
    define_asset_job,
    AssetSelection,
    ScheduleDefinition,
)
import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}


def get_page_content(url):
    response = requests.get(url, headers=headers)
    return response.content


@asset
def get_paths(context: AssetExecutionContext):
    paths = ["body-care"]

    return paths


@asset
def get_urls(context: AssetExecutionContext, get_paths: list):
    urls = []
    # Base URL of the website
    base_url = "https://www.bigc.co.th/category/{item_path}?page={page}"
    for path in get_paths:
        # Parse the HTML content
        soup_for_page = BeautifulSoup(
            get_page_content(base_url.format(item_path=path, page=1)), "lxml"
        )

        pagination = soup_for_page.find_all(
            "div", class_="pagination_pagination__wJ_sG"
        )

        last_page = re.findall('page=[0-9]*"', str(pagination[0]))[-1].split("=")[-1][
            :-1
        ]

        for i in range(1, int(last_page) + 1):
            urls.append(base_url.format(item_path=path, page=i))

    return urls


@asset
def get_page_contents(context: AssetExecutionContext, get_urls: list):
    response_contents = []
    for url in get_urls:
        response_contents.append(get_page_content(url))

    return response_contents


@asset
def process_bigc_data(context: AssetExecutionContext, get_page_contents: list):
    product_data = []
    for page_content in get_page_contents:
        soup = BeautifulSoup(page_content, "lxml")

        products = soup.find_all("div", class_="productCard_container__KXMQK")

        # Extract product names and prices
        for product in products:
            try:
                name_tag = product.find(
                    "div", class_="productCard_title__f1ohZ"
                ).text.strip()
                price_tag = product.find(
                    "div", class_="productCard_price__9T3J8"
                ).text.strip()
                price_tag = re.findall("[.0-9]+", price_tag)[0]
                product_data.append([name_tag, price_tag])
            except AttributeError:
                continue

    df = pd.DataFrame(product_data, columns=["Product Name", "Price"])

    return df


@asset
def beauty_excel(context: AssetExecutionContext, process_bigc_data: pd.DataFrame):

    wb = Workbook()
    ws = wb.active
    ws.title = "product_price"

    # Define header fill and font styles
    header_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    header_font = Font(bold=True)

    # Define border style
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for r_idx, row in enumerate(
        dataframe_to_rows(process_bigc_data, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font

            # Apply border to each cell
            cell.border = thin_border

    # Set the column width to a fixed width or based on the max length of content in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # filename with date and time
    today = pd.Timestamp.now().strftime("%Y%m%d_%H-%M-%S")
    filename = f"product_bigc_{today}.xlsx"
    wb.save("./data/" + filename)

    return filename


@asset
def upload_file_drive(context: AssetExecutionContext, beauty_excel: str):
    # Define the Google Drive API scopes and service account file path
    SCOPES = ["https://www.googleapis.com/auth/drive"]
    SERVICE_ACCOUNT_FILE = "./credential.json"

    destination_folder_id = "1BTKRtc1CxF4E1MQ7mOZFAHNfh4ni3Jq-"

    # Create credentials using the service account file
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES
    )

    # Build the Drive API service
    service = build("drive", "v3", credentials=credentials)

    # Upload the file to the destination folder
    file_metadata = {
        "name": beauty_excel,
        "parents": [destination_folder_id],  # Specify the destination folder ID
    }

    media = MediaFileUpload(
        "./data/" + beauty_excel,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    file = (
        service.files()
        .create(body=file_metadata, media_body=media, fields="id")
        .execute()
    )


defs = Definitions(
    assets=[
        get_paths,
        get_urls,
        get_page_contents,
        process_bigc_data,
        upload_file_drive,
        beauty_excel,
    ],
    jobs=[
        define_asset_job(
            name="process_bigc_data_and_upload", selection=AssetSelection.all()
        )
    ],
    schedules=[
        ScheduleDefinition(
            name="bigc_schedule",
            job_name="process_bigc_data_and_upload",
            cron_schedule="* * * * *",
        )
    ],
)
